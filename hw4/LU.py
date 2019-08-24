# -*- coding: utf-8 -*-
"""
Created on Wed Nov 28 00:26:06 2018

@author: user
"""
import openpyxl as pyxl
import numpy as np
import pprint
import scipy
import scipy.linalg   # SciPy Linear Algebra Library

def Horner(c,x,n):
    summ=c[n];
    for i in range(n-1,-1,-1):
        summ= summ*x+c[i];
    return summ;

def LUdecomp(a):
    n = len(a)
    for k in range(0,n-1):
        for i in range(k+1,n):
            if a[i,k] != 0.0:
                lam = a [i,k]/a[k,k]
                a[i,k+1:n] = a[i,k+1:n] - lam*a[k,k+1:n]
                a[i,k] = lam
    return a

def LUsolve(a,b):
    n = len(a)
    for k in range(1,n):
        b[k] = b[k] - np.dot(a[k,0:k],b[0:k])
    b[n-1] = b[n-1]/a[n-1,n-1]
    for k in range(n-2,-1,-1):
        b[k] = (b[k] - np.dot(a[k,k+1:n],b[k+1:n]))/a[k,k]
    return b

def back_sub(a,b,n):
    
    x=np.zeros(n,'float64')
    for i in range(n-1,-1,-1):
        x[i]=b[i]/a[i,i]
        for j in range(i-1,-1,-1):
            b[j]=b[j]-a[j,i]*x[i]
    #print("x:")
    #print(x)
    return x
        
def gauss(a,B,n):
    #A=np.copy(a)
    #b=np.copy(B)
    A=a
    b=B
    for i in range(0,n-1):
		# Partial pivoting
        maxEntry = abs(A[i][i])
        p = i
        for k in range(i,n):
            if(abs(A[k][i])>maxEntry):
                p = k
                maxEntry = abs(A[k][i])
                
        if(p!=i):
            for j in range(i,n):
                t = A[p][j]
                A[p][j] = A[i][j]
                A[i][j] = t
            t = b[p]
            b[p] = b[i]
            b[i] = t
		
		#Forward elimination.
        for k in range(i+1,n):
            if(A[k][i]==0.0):
                continue
            r = A[k][i]/A[i][i]
            for j in range(i,n):
                A[k][j] = A[k][j] - r*A[i][j]
            b[k] = b[k] - r*b[i]
    return a,b
	

def create_H_vec(A,i,j,v,n):
    for k in range(0,i):
        v[k]=0
    norm2 = 0.0
    for k in range(i,n):
        norm2 = norm2+A[k,j]*A[k,j]
    norm2 = norm2**0.5
    if(A[i,j]>=0):
        v[i] = A[i,j] + norm2
    else:
        v[i] = A[i,j] - norm2
    for k in range(i+1,n):
        v[k] = A[k,j]
    print(" Householder vector=")
    #for k in range(0,n):
    print(v)
    print("\n")

def QR_reflect(A,b,m,n):
    v=np.ones(m,'float64')
    #print(v)
    vv=0
    vx=0
    for j in range(0,n):
        create_H_vec(A,j,j,v,m)
        #print (v)
        vv=float(0.0)
        for i in range(0,m):
            vv += v[i]*v[i]
        for k in range(j,n):
            vx = float(0.0)
            for i in range(j,m):
                vx = vx+ A[i,k]*v[i]
            for i in range(j,m):
                A[i,k] = A[i,k]-2.0*(vx/vv)*v[i]
        vx=0
        for i in range(0,m):
            vx+=v[i]*b[i]
        for i in range(j,m):
            b[i]=b[i] - 2.0*(vx/vv)*v[i]
        #print(vv,vx)
        print("A[]=")
        print(A)
        print("b[]=")
        print(b)
    return A,b
        

def QR_solver(A,x,b,m,n):
    QR_A,QR_b=QR_reflect(A,b,m,n)
    #print("b[]=")
    #print(b)
    x=back_sub(QR_A,QR_b,min(m,n))
    #print("x[]=")
    #print(x)
    
    return x


        


n=7
y=np.zeros(n+4)
c=np.ones(n+4)
#generate test_data
for i in range(0,n+4):
    y[i]=Horner(c,1.0+i*0.2,n)
#generate origin system
A=np.ones((n+4,n+1),'float64')
for i in range(0,n+4):
    for j in range(1,n+1):
        v= 1+i*0.2
        temp=v
        for k in range(0,j-1):
            v=v*temp
        A[i][j]=v
#generate new system
B=(A.T).dot(A)
d=(A.T).dot(y)

P, L, U = scipy.linalg.lu(B)
Q_new, R_new = scipy.linalg.qr(B)
Q_ori, R_ori = scipy.linalg.qr(A)

x_QRnew=back_sub(R_new,(Q_new.T).dot(d),n+1)
x_QRori=back_sub(R_ori,(Q_ori.T).dot(y),n+1)

B_test=np.copy(B)
d_test = np.copy(d)
B_gauss,d_gauss = gauss(B_test,d_test,n+1)
x_gauss = back_sub(B_gauss,d_gauss,n+1)

B_copy=np.copy(B)
d_test = np.copy(d)
test = LUdecomp(B_copy)
x_lu = LUsolve(test,d_test)

A_QR = np.copy(A)
y_QR = np.copy(y)
x_QR = np.zeros(n+4)
#print(y_QR)
x_QR=QR_solver(A_QR,x_QR,y_QR,n+4,n+1)

B_QR_2 = np.copy(B)
d_QR_2 = np.copy(d)
x_QR_2 = np.zeros(n+1)
x_QRnolib_new=QR_solver(B_QR_2,x_QR_2,d_QR_2,n+1,n+1)

'''print("After QR decomposition, A[][]=")
print(B_QR_2)
print("The solution x[]=")
for i in range(0,n+1):
    print("%.10lf"%x_QRnolib_new[i])'''


#output to xl
#第1頁是基本的
wb = pyxl.Workbook()
ws1 = wb.worksheets[0]
ws1.title = 'result'
ws1.append(['x_QRnew'])
ws1.append(x_QRnew.tolist())
ws1.append(['x_QRori'])
ws1.append(x_QRori.tolist())
ws1.append(['x_gauss'])
ws1.append(x_gauss.tolist())
ws1.append(['x_lu'])
ws1.append(x_lu.tolist())
ws1.append(['沒套函示庫QR_ori'])
ws1.append(x_QR.tolist())
ws1.append(['沒套函示庫QR_new'])
ws1.append(x_QRnolib_new.tolist())
#第2頁是QR 2~15
#第3頁是QR 2-norm 
#第4頁是 infinte-norm
n=2
ws2 = wb.create_sheet()
ws2.title = 'QR 2~15'
ws2.append(['沒套函示庫QR_ori'])
ws3 = wb.create_sheet()
ws3.title = '2-norm'
ws3.append(['沒套函示庫QR_ori'])
ws3.append(['2-norm'])
ws4 = wb.create_sheet()
ws4.title = 'infinite-norm'
ws4.append(['沒套函示庫QR_ori'])
ws4.append(['infinite-norm'])
while(n<15):
    A=np.ones((n+4,n+1),'float64')
    y=np.zeros(n+4)
    c=np.ones(n+4)
    for i in range(0,n+4):
        for j in range(1,n+1):
            v= 1+i*0.2
            temp=v
            for k in range(0,j-1):
                v=v*temp
            A[i][j]=v
    for i in range(0,n+4):
        y[i]=Horner(c,1.0+i*0.2,n)
    A_QR = np.copy(A)
    y_QR = np.copy(y)
    x_QR = np.zeros(n+4)
    #print(y_QR)
    x_QR=QR_solver(A_QR,x_QR,y_QR,n+4,n+1)
    ws2.append([n])
    ws2.append(x_QR.tolist())
    #ws3.append([n])
    ws3.append([(((x_QR-1)**2).sum())**0.5])
    #ws4.append([n])
    ws4.append([np.max(np.absolute(x_QR-1))])
    n=n+1

wb.save('result.xlsx')
